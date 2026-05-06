import argparse
import csv
import html
import os
import re
import shutil
import time
import uuid
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin
from urllib.request import Request, urlopen

from openpyxl import load_workbook


class LinkParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.hrefs = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() != "a":
            return

        for name, value in attrs:
            if name.lower() == "href" and value:
                self.hrefs.append(value)


def clean_file_name(name):
    if not name or not str(name).strip():
        return "untitled"

    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip()
    return cleaned[:120].strip() or "untitled"


def fetch_bytes(url, timeout):
    request = Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0 Safari/537.36"
            )
        },
    )
    with urlopen(request, timeout=timeout) as response:
        return response.read()


def find_pdf_links(page_url):
    body = fetch_bytes(page_url, timeout=30).decode("utf-8", errors="replace")
    parser = LinkParser()
    parser.feed(body)

    matches = []
    seen = set()
    for href in parser.hrefs:
        decoded = html.unescape(href)
        if "digidoc.xhtml" not in decoded or "downloadAttachment" not in decoded:
            continue

        absolute = urljoin(page_url, decoded)
        if absolute not in seen:
            seen.add(absolute)
            matches.append(absolute)

    return matches


def validate_pdf(path):
    size = path.stat().st_size
    if size < 1024:
        return False, "File is too small to be a valid PDF"

    with path.open("rb") as file:
        header = file.read(5)
        if header != b"%PDF-":
            return False, "File does not start with the PDF signature"

        tail_length = min(4096, size)
        file.seek(-tail_length, os.SEEK_END)
        tail = file.read(tail_length)
        if b"%%EOF" not in tail:
            return False, "File does not contain a PDF EOF marker near the end"

    return True, ""


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    parser = argparse.ArgumentParser(
        description="Download and validate WOAH PDFs listed in the DATABASE worksheet."
    )
    parser.add_argument(
        "--workbook-path",
        default="./Database (5 mai 2026)/WOAH-DOI-database-v4_(CopyForDIAD).xlsx",
        help="Path to the Excel workbook.",
    )
    parser.add_argument(
        "--output-dir",
        default="./downloads-test",
        help="Directory where PDFs and the CSV report are written.",
    )
    parser.add_argument("--start-row", type=int, default=3295)
    parser.add_argument("--end-row", type=int, default=3738)
    args = parser.parse_args()

    workbook_path = Path(args.workbook_path)
    output_dir = Path(args.output_dir)
    clean_pdf_dir = output_dir / "clean-pdfs"
    temp_dir = output_dir / "_temp"
    report_path = output_dir / "download_report.csv"

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    clean_pdf_dir.mkdir(parents=True, exist_ok=True)
    temp_dir.mkdir(parents=True, exist_ok=True)

    rows = []

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        worksheet = workbook["DATABASE"]
        last_row = min(worksheet.max_row, args.end_row)

        for row_number in range(args.start_row, last_row + 1):
            rank = cell_text(worksheet.cell(row=row_number, column=1).value)
            title = cell_text(worksheet.cell(row=row_number, column=7).value)
            doi = cell_text(worksheet.cell(row=row_number, column=12).value)
            record_url = cell_text(worksheet.cell(row=row_number, column=13).value) or doi

            if not record_url:
                print(f"WARNING: Row {row_number} has no URL or DOI. Skipping.")
                rows.append(
                    {
                        "Row": row_number,
                        "Rank": rank,
                        "Title": title,
                        "DOI": doi,
                        "RecordUrl": record_url,
                        "DownloadUrl": "",
                        "OutputFile": "",
                        "Status": "MissingUrl",
                        "Issue": "No URL or DOI in row",
                        "Bytes": 0,
                    }
                )
                continue

            print(f"Row {row_number}: opening {record_url}")

            try:
                pdf_links = find_pdf_links(record_url)
                if not pdf_links:
                    print(f"WARNING: Row {row_number}: no PDF download link found.")
                    rows.append(
                        {
                            "Row": row_number,
                            "Rank": rank,
                            "Title": title,
                            "DOI": doi,
                            "RecordUrl": record_url,
                            "DownloadUrl": "",
                            "OutputFile": "",
                            "Status": "NoPdfLink",
                            "Issue": "No digidoc PDF download link found on portal page",
                            "Bytes": 0,
                        }
                    )
                    continue

                for index, download_url in enumerate(pdf_links, start=1):
                    file_stem = clean_file_name(f"{rank} - {title}")
                    if len(pdf_links) > 1:
                        file_stem = f"{file_stem} - {index}"

                    target_file = clean_pdf_dir / f"{file_stem}.pdf"
                    temp_file = temp_dir / f"{uuid.uuid4()}.pdf"

                    print(f"Row {row_number}: downloading and validating {target_file}")
                    temp_file.write_bytes(fetch_bytes(download_url, timeout=60))

                    is_valid, issue = validate_pdf(temp_file)
                    bytes_count = temp_file.stat().st_size

                    if is_valid:
                        shutil.move(str(temp_file), str(target_file))
                        rows.append(
                            {
                                "Row": row_number,
                                "Rank": rank,
                                "Title": title,
                                "DOI": doi,
                                "RecordUrl": record_url,
                                "DownloadUrl": download_url,
                                "OutputFile": str(target_file),
                                "Status": "Clean",
                                "Issue": "",
                                "Bytes": bytes_count,
                            }
                        )
                        print(f"Row {row_number}: clean PDF saved")
                    else:
                        temp_file.unlink(missing_ok=True)
                        rows.append(
                            {
                                "Row": row_number,
                                "Rank": rank,
                                "Title": title,
                                "DOI": doi,
                                "RecordUrl": record_url,
                                "DownloadUrl": download_url,
                                "OutputFile": "",
                                "Status": "CorruptedPdf",
                                "Issue": issue,
                                "Bytes": bytes_count,
                            }
                        )
                        print(f"WARNING: Row {row_number}: corrupted PDF discarded - {issue}")

                    time.sleep(0.5)
            except Exception as exc:
                print(f"WARNING: Row {row_number} failed: {exc}")
                rows.append(
                    {
                        "Row": row_number,
                        "Rank": rank,
                        "Title": title,
                        "DOI": doi,
                        "RecordUrl": record_url,
                        "DownloadUrl": "",
                        "OutputFile": "",
                        "Status": "Failed",
                        "Issue": str(exc),
                        "Bytes": 0,
                    }
                )
    finally:
        if rows:
            with report_path.open("w", newline="", encoding="utf-8") as file:
                writer = csv.DictWriter(
                    file,
                    fieldnames=[
                        "Row",
                        "Rank",
                        "Title",
                        "DOI",
                        "RecordUrl",
                        "DownloadUrl",
                        "OutputFile",
                        "Status",
                        "Issue",
                        "Bytes",
                    ],
                )
                writer.writeheader()
                writer.writerows(rows)

            print(f"Report written to {report_path}")
            print(f"Clean PDFs are in {clean_pdf_dir}")

        shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    main()
